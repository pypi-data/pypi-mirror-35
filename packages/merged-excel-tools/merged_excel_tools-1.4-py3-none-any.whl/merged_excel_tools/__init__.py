class merged_excel():
    ws = None

    def __init__(self, xls_name):
        from openpyxl import load_workbook
        wb = load_workbook(filename=xls_name)
        self.ws = wb.active

    def getMergedList(self):
        ranges = self.ws.merged_cells.ranges
        myRanges = [x for x in ranges if x.min_col == 1 and x.max_col == 1]
        myRanges.sort(key=lambda x: x.min_row)
        return myRanges

    def getValue(self, area, index):
        return self.ws[("%s%d" % (area, index))].value
