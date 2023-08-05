class merged_excel():
    ws = None

    def __init__(self, xls_name):
        from openpyxl import load_workbook
        wb = load_workbook(filename=xls_name)
        self.ws = wb.active

    def getMergedList(self):
        import re
        ranges = self.ws.merged_cell_ranges
        myRanges = [x for x in ranges if re.match("A[0-9]+:A[0-9]+", x)]
        myRanges.sort(key=lambda x: self.getRowFromRange(x, 0))
        return myRanges

    def getRowFromRange(self, range, index):
        return int(range.split(':')[index][1:])

    def get(self,area,index):
        return self.ws[("%s%d" % (area,index))].value

    def get(self,name):
        return self.ws[name].value
