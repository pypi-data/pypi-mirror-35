__author__ = 'Kevin'

from basedriver import *
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from .. import helper
import os

class OpenpyxlDriver(BaseDriver):
    """ Abstract base-driver """
    def __init__(self):
        """ Constructor """
        super(OpenpyxlDriver, self).__init__()

    def open_workbook_from_file(self, filename):
        """

        :param filename: Path af File
        :type filename: str
        :rtype: BaseWorkbook
        """
        wb = load_workbook(filename, data_only=True, keep_vba=True)
        return OpenpyxlWorkbook(self, wb, os.path.basename(filename))

    def open_workbook(self):
        """

        :rtype: BaseWorkbook
        """
        wb = Workbook()
        return OpenpyxlWorkbook(self, wb)

class OpenpyxlWorkbook(BaseWorkbook):
    """ Abstract Workbook """
    def __init__(self, driver, workbook, name='NONAME'):
        """ Constructor """
        super(OpenpyxlWorkbook, self).__init__(driver)
        self.workbook = workbook
        self._name = name

    def get_sheet(self, name):
        """

        :param name: Name of Sheet
        :type name: str
        :return: Sheet
        :rtype: BaseWorksheet
        """
        return OpenpyxlWorksheet(self, self.workbook[name])

    @property
    def sheets(self):
        """

        :return: Tuple of containing sheets
        :rtype: tuple of [BaseWorksheet]
        """
        return [OpenpyxlWorksheet(self, self.workbook[x]) for x in self.workbook.sheetnames]

    @property
    def name(self):
        """

        :return: Name of Workbook
        :rtype: str
        """
        return self._name

    @property
    def path(self):
        """

        :return: Path of Workbook
        :rtype: str
        """
        raise NotImplementedError()

    def save(self):
        """
        Saves the Workbook to Disk.
        """
        raise NotImplementedError()

    def save_as(self, filepath):
        """
        Saves the Workbook under given Path to Disk.
        :param filepath: Path to save Workbook
        :type filepath: str
        """
        self.workbook.save(filepath)
        return self

    def close(self):
        """
        Close the Workbook
        """
        pass

class OpenpyxlWorksheet(BaseWorksheet):
    """ Abstract Worksheet """
    def __init__(self, workbook, worksheet):
        """ Constructor """
        super(OpenpyxlWorksheet, self).__init__(workbook)
        self.worksheet = worksheet

    @property
    def name(self):
        """

        :return: Name of Sheet
        :rtype: str
        """
        return self.worksheet.title

    @name.setter
    def name(self, value):
        """
        Sets the Name of the Sheet
        :param value: New name of the Sheet
        :type value: str
        """
        self.worksheet.title = value

    def range(self, start, end):
        """

        :param start: Start position in 2-Dimensions (f. example: (1, 42))
        :type start: tuple of [int]
        :param end: End position in 2-Dimensions (f. example: (1, 42)
        :type end: tuple of [int]
        :return: Range Object
        :rtype: BaseRange
        """

        return OpenpyxlRange(self, start, end)

    def cell(self, position):
        """

        :param position: Position in 2-Dimensions (f. example: (1, 42))
        :type position: tuple of [int]
        :return: Cell Object
        :rtype: BaseCell
        """
        return OpenpyxlCell(self, position,
                            self.worksheet.cell(
                                row=position[0],
                                column=position[1]
                            )
        )

    def addImage(self, position, imagePath):
        """
        :param position: Position in 2-Dimensions (f. example: (1, 42))
        :type position: tuple of [int]
        :param imagePath: Path to image file
        :type position: str
        """
        address = helper.pos2address(position[0], position[1])

        img = Image(imagePath)
        img.anchor = address

        self.worksheet.add_image(img)


class OpenpyxlRange(BaseRange):
    """ Abstract Range """
    def __init__(self, sheet, start, end):
        """ Constructor """
        super(OpenpyxlRange, self).__init__(sheet, start, end)

    def cell(self, position):
        """

        :param position: Position in 2-Dimensions (f. example: (1, 42))
        :type position: tuple of [int]
        :return: Cell Object
        :rtype: BaseCell
        """
        return self.worksheet.cell(
                                row=position[0] + self.start[0],
                                column=position[1] + self.start[1]
                            )

    @property
    def size(self):
        """

        :return: Size of Range in 2-Dimensions (f. example: (1, 42))
        :rtype: tuple of [int]
        """
        i, j = self.end[0] - self.start[0] + 1, self.end[1] - self.start[1] + 1
        return (i, j)

    @property
    def value(self):
        """

        :return: Values of all Cells containing in the in 2-Dimensions
        :rtype: tuple of [tuple of [object]]
        """
        v = []
        size = self.size
        for r in range(size[0]):
            r1 = r + self.start[0]
            v1 = []
            for c in range(size[1]):
                c1 = c + self.start[1]
                v1.append(self.worksheet.cell((r1,c1)).value)
            v.append(v1)
        return helper.reduce_dimensions(v, *size)


    @value.setter
    def value(self, value):
        """

        :param value: Values of all Cells containing in a in 2-Dimensions
        :type value: tuple of [tuple of [object]]
        """
        #self.range.Value2 = value
        for row in range(len(value)):
            value2 = value[row]
            r = row + self.start[0]
            for col in range(len(value2)):
                c = col + self.start[1]
                self.worksheet.cell((r,c)).value = value2[col]

    @property
    def formula(self):
        """

        :return: Formulas of all Cells containing in a in 2-Dimensions
        :rtype: tuple of [tuple of [str]]
        """
        v = []
        size = self.size
        for r in range(size[0]):
            r1 = r + self.start[0]
            v1 = []
            for c in range(size[1]):
                c1 = c + self.start[1]
                v1.append(self.worksheet.cell((r1,c1)).cell.internal_value)
            v.append(v1)
        return helper.reduce_dimensions(v, *size)

    @formula.setter
    def formula(self, value):
        """

        :param value: Formulas of all Cells containing in a in 2-Dimensions
        :type value: tuple of [tuple of [str]]
        """
        #self.range.Value2 = value
        for row in range(len(value)):
            value2 = value[row]
            r = row + self.start[0]
            for col in range(len(value2)):
                c = col + self.start[1]
                self.worksheet.cell(row=r, column=c).value = value2[col]

    def resize(self, heigth=0, width=0):
        """
        Resizes the Range by given Offsets
        :param heigth: Height offset
        :type heigth: int
        :param width: Width offset
        :type width: int
        :return: New, resized Range
        :rtype: BaseRange
        """
        raise NotImplementedError()

class OpenpyxlCell(BaseCell):
    """ Abstract Cell """
    def __init__(self, sheet, position, cell):
        """ Constructor """
        super(OpenpyxlCell, self).__init__(sheet, position)
        self.cell = cell

    @property
    def value(self):
        """

        :return: Value inside Cell
        :rtype: object
        """
        return self.cell.internal_value


    @value.setter
    def value(self, value):
        """
        Sets the value inside the Cell
        :param value: New Value
        :type value: object
        """
        self.cell.value = value

    @property
    def formula(self):
        """

        :return: Formula inside the Cell
        :rtype: str
        """
        return self.cell.value

    @formula.setter
    def formula(self, value):
        """

        :param value: Formula inside the Cell
        :type value: str
        """
        self.cell.value = value

    @property
    def comment(self):
        """

        :return: Comment of Cell
        :rtype: str
        """
        raise NotImplementedError()

    @comment.setter
    def comment(self, value):
        """
        Sets the comment of the Cell
        :param value: Comment of the Cell
        :type value: str
        """
        raise NotImplementedError()