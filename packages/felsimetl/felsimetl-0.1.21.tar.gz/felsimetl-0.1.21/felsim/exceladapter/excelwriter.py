import openpyxl


class ExcelWriter:
    def __init__(self, filename):
        self.workbook = openpyxl.Workbook()
        self.filename = filename

    def create_sheet(self, sheetname):
        return self.workbook.create_sheet(index=0, title=sheetname)

    def save(self):
        self.workbook.save(filename=self.filename)
