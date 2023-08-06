from felsim.constants import *
import openpyxl


class CategoriesReader:
    def __init__(self, filename):
        # sacar filename de acá
        rubros_wb = openpyxl.load_workbook(filename=filename, data_only=True)
        rubros_sheet = rubros_wb.get_sheet_by_name(RUBROS_SHEET_NAME)

        self.rubros = {}

        for rowNum in range(2, rubros_sheet.max_row + 1):  # skip the first row
            item = {}
            key = rubros_sheet.cell(row=rowNum, column=1).value
            item['category'] = rubros_sheet.cell(row=rowNum, column=2).value
            item['account'] = rubros_sheet.cell(row=rowNum, column=3).value
            item['details'] = rubros_sheet.cell(row=rowNum, column=4).value

            self.rubros[key] = item


    def get_category_from_details(self, details):
        account = ""
        category = ""

        for key, value in self.rubros.items():
            if value['details'] == details:
                account = value['account']
                category = value['category']
                break

        return [category, account]

    def get_category_from_account_and_details(self, account, details):
        if details.startswith(MORATORIA_AFIP):
            return IMPUESTOS

        key = "%s-%s" % (account, details)
        if key in self.rubros:
            return self.rubros[key]['category']
        else:
            return NEW_CATEGORY
