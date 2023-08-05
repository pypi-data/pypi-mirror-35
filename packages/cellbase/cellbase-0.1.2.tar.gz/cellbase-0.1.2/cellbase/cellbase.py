import os

from openpyxl import Workbook, load_workbook

from cellbase.helper import CellFormatter
from cellbase.celltable import Celltable


class Cellbase:
    """
    Cellbase is equivalent to :class:`Workbook` which stores :class:`Celltable`
    """
    DEFAULT_FILENAME = 'cellbase.xlsx'

    def __init__(self):
        self.filename = os.path.join(os.getcwd(), Cellbase.DEFAULT_FILENAME)
        self.on_create = {}
        self.workbook = Workbook()
        self.celltables = {}

    def load(self, filename):
        """
        Load workbook from given filename

        Notice how this named as "load" instead of "open" as it does not open a connection or stream with the workbook.
        Instead, it simply load the data into memory and any changes will only be saved unless save or save_as
        is called.

        :param filename: Path of workbook to load
        :type filename: str
        :return: self
        :rtype: Cellbase
        """
        self.filename = filename
        self.workbook = load_workbook(filename) if os.path.exists(filename) else Workbook()
        for worksheet in self.workbook.worksheets:
            self.celltables[worksheet.title] = Celltable(worksheet)
        return self

    def remove_empty_cols(self, worksheet_name):
        """
        Remove 1st row's columns where its value is None. It does not inspect the whole column, so use it with care if
        there's data that you would like to preserve under empty header

        :param worksheet_name: Name of worksheet to remove empty columns
        :type worksheet_name str
        """
        worksheet = self.workbook[worksheet_name]
        for empty_col in reversed([col_id for col_id in worksheet[1] if col_id.value is None]):
            worksheet.delete_cols(empty_col.col_idx)

    def register(self, on_create):
        """
        Register format of worksheet to deal with, only required for newly created worksheet
        Example::

        {'TABLE_NAME': ['COL_NAME_1', 'COL_NAME_2']}
        :param on_create: Format of Celltable to deal with
        :type on_create: dict
        :return:
        """
        self.on_create.update(on_create)

    def create_if_none(self, worksheet_name):
        """
        Create worksheet and add to cell_tables if there's no such worksheet. It is first called in every data
        accessing methods like query, insert, update, etc.

        :param worksheet_name: Name of worksheet to inspect or create if required
        :type worksheet_name: str
        """
        if worksheet_name not in self.celltables:
            if self.on_create is None:
                raise ValueError(
                    "Trying to create Celltable '%s' without specifying details in on_create" % worksheet_name)
            worksheet = self.workbook.create_sheet(title=worksheet_name)
            worksheet.append(self.on_create[worksheet_name])
            self.celltables[worksheet.title] = Celltable(worksheet)

    def query(self, worksheet_name, where=None):
        """
        Return data from Celltable with specified worksheet_name, that match the conditions.
        Return all data if no condition given.

        :param worksheet_name: Name of worksheet to query from
        :type worksheet_name: str
        :param where: dict of columns id to inspect. For example, {'id': 1, 'name': 'jp'}.
        :type where: dict
        :return:
            List of dict that store value corresponding to the column id.
            * row_idx is the default value to return, where it specifies the row index of returned data.
            row_idx is corresponding to the actual row index in spreadsheet, so the minimum index is 2 where 1st row
            is taken by the column ids(header)
            For example, [{"row_idx": 2, "id": 1, "name": "jp1"}, {"row_idx": 3, "id": 2, "name": "jp2"}]
        :rtype: list
        """
        self.create_if_none(worksheet_name)
        return self.celltables[worksheet_name].query(where=where)

    def insert(self, worksheet_name, value_in_dict):
        """
        Insert new row to the worksheet

        :param worksheet_name: Name of worksheet to insert to
        :type worksheet_name: str
        :param value_in_dict:
            Dict that describe the row to insert, where row_idx is not required.
            For example, {"id": 1, "name": "jp1"}
        :type value_in_dict: dict
        :return: row_idx of new row
        :rtype: int
        """
        self.create_if_none(worksheet_name)
        return self.celltables[worksheet_name].insert(value_in_dict)

    def update(self, worksheet_name, value_in_dict, where=None):
        """
        Update row(s) that match the condition.
        If row_idx is given in value_in_dict, wheres & conds will be ignored and only the exact row will be updated.

        :param worksheet_name: Name of the worksheet to update
        :type worksheet_name: str
        :param value_in_dict: Dict that describe the row, where row_idx is optional.
        :type value_in_dict: dict
        :param where: dict of columns id to inspect. For example, {'id': 1, 'name': 'jp'}.
        :type where: dict
        :return: Number of rows updated
        :rtype: int
        """
        self.create_if_none(worksheet_name)
        return self.celltables[worksheet_name].update(value_in_dict, where=where)

    def delete(self, worksheet_name, where=None):
        """
        Delete row(s) that match conditions.

        :param worksheet_name: Name of worksheet to delete row(s)
        :type worksheet_name: str
        :param where: dict of columns id to inspect. For example, {'id': 1, 'name': 'jp'}.
        :type where: dict
        :return: Number of rows deleted
        :rtype: int
        """
        self.create_if_none(worksheet_name)
        return self.celltables[worksheet_name].delete(where=where)

    def traverse(self, worksheet_name, fn, where=None, select=None):
        """
        Access cells directly from rows where conditions matched.

        :param worksheet_name: Name of worksheet to traverse
        :type worksheet_name: str
        :param fn:
            function(:class:`openpyxl.cell.Cell`) to allow accessing the cell.
            For example, lambda cell: cell.fill = PatternFill(fill_type="solid", fgColor="00FFFF00").
        :param where: dict of columns id to inspect. For example, {'id': 1, 'name': 'jp'}.
        :type where: dict
        :param select:
            The columns of the row to update.
            For example, ["id"], where only column under "id" will be accessed
        :type select: list
        :return: Number of rows traversed
        :rtype: int
        """
        self.create_if_none(worksheet_name)
        return self.celltables[worksheet_name].traverse(fn, where=where, select=select)

    def format(self, worksheet_name, where=None, select=None,
               formatter=None,
               font=None, fill=None, border=None, number_format=None,
               protection=None, alignment=None, style=None
               ):
        """
        Convenience method that built on top of traverse to format cell(s).
        If formatter is given, all other formats will be ignored.

        :param worksheet_name: Name of worksheet to format
        :type worksheet_name: str
        :param where: dict of columns id to inspect. For example, {'id': 1, 'name': 'jp'}.
        :type where: dict
        :param select:
            The columns of the row to update.
            For example, ["id"], where only column under "id" will be formatted
        :type select: list
        :param formatter:
            CellFormatter that hold all formats. When this is not None other formats will be ignored.
        :type formatter: CellFormatter
        :param font: Font of cell
        :type font: openpyxl.styles.Font
        :param fill: Fill cell with color
        :type fill: openpyxl.styles.Fill
        :param border: Border of cell
        :type border: openpyxl.styles.Border
        :param number_format: Number format of cell
        :type number_format: str
        :param protection: Protection of cell
        :type protection: openpyxl.styles.Protection
        :param alignment: Alignment of cell
        :type alignment: openpyxl.styles.Alignment
        :param style: Named style
        :type style: str
        :return: Number of rows formatted
        :rtype: int
        """
        self.create_if_none(worksheet_name)
        if formatter is None:
            formatter = CellFormatter(
                font=font, fill=fill, border=border,
                number_format=number_format, protection=protection,
                alignment=alignment, style=style
            )
        return self.celltables[worksheet_name].format(
            formatter, where=where, select=select
        )

    def drop(self, worksheet_name):
        """
        Delete specified worksheet.

        :param worksheet_name: Name of worksheet to delete
        :type worksheet_name: str
        """
        worksheet_to_drop = self.celltables[worksheet_name].worksheet
        # Workbook must contain at least 1 visible sheet
        visible_sheets = [worksheet for worksheet in self.workbook.worksheets
                          if worksheet.sheet_state == 'visible']
        if len(visible_sheets) == 1 and visible_sheets[0] is worksheet_to_drop:
            self.workbook.create_sheet()
        self.workbook.remove(worksheet_to_drop)
        self.celltables.pop(worksheet_name)

    def save(self):
        """
        Save workbook to the filename specified in open, overwrite if file exist.
        """
        self.save_as(self.filename, overwrite=True)

    def save_as(self, filename, overwrite=False):
        """
        Save workbook to filename. FileExistsError will be raised if file exists and overwrite is False.

        :param filename: Path to save the workbook
        :type filename: str
        :param overwrite: Whether to overwrite if file exists
        :type overwrite: bool
        :raises FileExitsError: File exists and overwrite is False
        """
        if os.path.exists(filename) and overwrite is False:
            raise FileExistsError("%s already exists, set overwrite=True if this is expected.")
        self.workbook.save(filename)

    def __len__(self):
        """
        Return numbers of worksheet

        :return: Numbers of worksheet
        """
        return len(self.celltables)

    def __getitem__(self, worksheet_name):
        """
        Get Celltable with worksheet_name

        :param worksheet_name: worksheet_name to find
        :return: Celltable to find
        :rtype: Celltable
        """
        self.create_if_none(worksheet_name)
        return self.celltables[worksheet_name]

    def __setitem__(self, worksheet_name, celltable):
        """
        Celltable must be created by Cellbase

        :raise AssertionError: When attempt to assign
        """
        raise AssertionError("Celltable must be created by Cellbase")

    def __delitem__(self, worksheet_name):
        """
        Drop worksheet

        :param worksheet_name: Name of worksheet to drop
        :type worksheet_name: str
        """
        self.drop(worksheet_name)

    def __contains__(self, worksheet_name):
        """
        Check if worksheet exists

        :param worksheet_name: Name of worksheet to check
        :type worksheet_name: str
        :return: If worksheet exists
        :rtype: bool
        """
        return worksheet_name in self.celltables
