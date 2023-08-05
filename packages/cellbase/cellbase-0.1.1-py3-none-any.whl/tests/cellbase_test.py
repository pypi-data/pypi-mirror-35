import unittest

from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, Protection
from openpyxl.styles.numbers import FORMAT_TEXT
from cellbase import Cellbase, DAO, Entity, CellFormatter
from cellbase.celltable import Celltable


class CellbaseTest(unittest.TestCase):
    def setUp(self):
        self.cellbase = Cellbase().load("../out/not_exist.xlsx")
        self.cellbase.register(on_create=SimpleDAO.on_create())
        self.dao = SimpleDAO(self.cellbase)
        self.assertEqual(Simple(id=1, name="name"), Simple(id=1, name="name"))  # Test entity's equality

    def test_insert_and_query(self):
        simple = Simple(id=1, name="test_simple")
        self.dao.insert(simple)
        self.assertTrue(simple.row_idx in self.cellbase.celltables[SimpleDAO.TABLE_NAME])
        self.assertEqual(simple, self.dao.query({DAO.COL_ROW_IDX: lambda row_idx: row_idx > 1})[0])

    def test_query_all_and_row_idx_lambda(self):
        for i in range(5):
            self.dao.insert(Simple(id=i, name="simple%s" % i))
        row_size = len(self.cellbase.celltables[SimpleDAO.TABLE_NAME].rows)
        self.assertEqual(row_size, len(self.dao.query()))
        # 5 rows inserted to row index 2, 3, 4, 5, 6
        simples_row_idx_4_to_6 = self.dao[lambda row_idx: 4 <= row_idx <= 6]
        self.assertEqual(len(simples_row_idx_4_to_6), 3)
        for i in range(3):
            self.assertEqual(simples_row_idx_4_to_6[i].row_idx, i + 4)

    def test_update(self):
        simple = Simple(id=2, name="orig_simple")
        self.dao.insert(simple)
        simple_to_update = Simple(id=3, name="updated_simple")
        simple_to_update.row_idx = simple.row_idx
        self.assertNotEqual(simple_to_update, self.dao.query({DAO.COL_ROW_IDX: simple.row_idx})[0])
        self.dao.update(simple_to_update, {DAO.COL_ROW_IDX: 2})
        self.assertEqual(simple_to_update, self.dao.query({DAO.COL_ROW_IDX: simple.row_idx})[0])
        self.assertNotEqual(simple, simple_to_update)

    def test_delete(self):
        simple = Simple(id=4, name="simple_to_delete")
        self.dao.insert(simple)
        deleted_count = self.dao.delete({DAO.COL_ROW_IDX: simple.row_idx})
        self.assertEqual(deleted_count, 1)
        self.assertEqual(self.dao.query({DAO.COL_ROW_IDX: simple.row_idx}), [])

    def test_format(self):
        simple_formatted = Simple(id=5, name="simple_formatted")
        simple_not_formatted = Simple(id=6, name="simple_not_formatted")
        self.dao.insert(simple_formatted)
        self.dao.insert(simple_not_formatted)
        font = Font(name='Arial')
        fill = PatternFill(fill_type="solid", fgColor="00FFFF00")
        border = Border(top=Side(style="thin"))
        alignment = Alignment(horizontal="left")
        number_format = FORMAT_TEXT
        protection = Protection(hidden=True)
        formatter = CellFormatter(
            font=font, fill=fill, border=border,
            number_format=number_format, protection=protection,
            alignment=alignment)
        self.dao.format({DAO.COL_ROW_IDX: simple_formatted.row_idx},
                        font=font, fill=fill, border=border,
                        number_format=number_format, protection=protection,
                        alignment=alignment)
        self.dao.traverse(lambda cell:
                          self.assertTrue(
                              is_all_format_match(cell, formatter)
                          ), {DAO.COL_ROW_IDX: simple_formatted.row_idx})
        self.dao.traverse(lambda cell:
                          self.assertFalse(
                              is_all_format_match(cell, formatter)
                          ), {DAO.COL_ROW_IDX: simple_not_formatted.row_idx})

    def test_dao_and_celltable_magic_methods(self):
        simple = Simple(id=1, name="test_simple")
        self.dao.insert(simple)
        self.assertTrue(simple.row_idx in self.dao)  # __contains__
        self.assertEqual(simple, self.dao[simple.row_idx])  # __getitem__ not callable return entity
        self.assertEqual(simple, self.dao[lambda row_idx: row_idx == simple.row_idx][0])  # callable return list
        self.assertEqual(len(self.dao), 1)  # __len__
        # __setitem__
        simple.name = "updated_simple"
        self.dao[simple.row_idx] = simple
        self.assertEqual(self.dao[simple.row_idx].name, "updated_simple")
        # __delitem__
        self.assertTrue(simple.row_idx in self.dao)
        self.dao.delete({DAO.COL_ROW_IDX: simple.row_idx})
        self.assertTrue(simple.row_idx not in self.dao)
        # _setitem__ with callable
        simple_new = Simple(id=0, name='new_simple')
        with self.assertWarns(UserWarning):
            self.dao[lambda row_idx: 2 <= row_idx <= 6] = simple_new
        self.assertEqual(len(self.dao), 0)  # Insert with callable should have no effect at all
        for i in range(5):  # Add row 2, 3, 4, 5, 6
            self.dao.insert(Simple(id=i, name="simple%s" % i))
        self.dao[lambda row_idx: 4 <= row_idx <= 6] = simple_new
        for i in range(3):
            simple_at = self.dao[i + 4]
            self.assertEqual(simple_new.id, simple_at.id)
            self.assertEqual(simple_new.name, simple_at.name)
        # __delitem__ with callable
        del self.dao[lambda row_idx: row_idx > 1]
        self.assertEqual(len(self.dao), 0)

    def test_cellbase_magic_methods(self):
        self.cellbase.create_if_none(SimpleDAO.TABLE_NAME)
        self.assertTrue(len(self.cellbase), 1)  # __len__
        self.assertTrue(SimpleDAO.TABLE_NAME in self.cellbase)  # __contains__
        self.assertEqual(self.cellbase[SimpleDAO.TABLE_NAME],
                         self.cellbase.celltables[SimpleDAO.TABLE_NAME])  # __getitem__
        # __delitem__
        del self.cellbase.celltables[SimpleDAO.TABLE_NAME]
        self.assertTrue(SimpleDAO.TABLE_NAME not in self.cellbase)
        # __setitem__
        with self.assertRaises(AssertionError):
            self.cellbase[SimpleDAO.TABLE_NAME] = Celltable(
                self.cellbase.workbook.create_sheet(title=SimpleDAO.TABLE_NAME))


def is_all_format_match(cell, formatter):
    return all([formatter.font.name == cell.font.name,
                formatter.fill.fgColor == cell.fill.fgColor,
                formatter.border.top.style == cell.border.top.style,
                formatter.alignment.horizontal == cell.alignment.horizontal,
                formatter.number_format == cell.number_format,
                formatter.protection.hidden == cell.protection.hidden])


class Simple(Entity):
    def __init__(self, id=0, name="simple"):
        super().__init__()
        self.id = id
        self.name = name

    def from_dict(self, values):
        super().from_dict(values)
        self.id = values[SimpleDAO.COL_ID]
        self.name = values[SimpleDAO.COL_NAME]
        return self
    
    def to_dict(self):
        values = super().to_dict()
        values.update({SimpleDAO.COL_ID: self.id, SimpleDAO.COL_NAME: self.name})
        return values

    def __eq__(self, other):
        if isinstance(other, Simple):
            return self.row_idx == other.row_idx and self.id == other.id and self.name == other.name
        return NotImplemented


class SimpleDAO(DAO):
    TABLE_NAME = "Simple"
    COL_ID = "id"
    COL_NAME = "name"

    def worksheet_name(self):
        return SimpleDAO.TABLE_NAME

    def new_entity(self):
        return Simple()

    @staticmethod
    def on_create():
        return {SimpleDAO.TABLE_NAME: [SimpleDAO.COL_ID, SimpleDAO.COL_NAME]}


if __name__ == "__main__":
    unittest.main()
